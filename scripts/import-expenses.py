"""
Import MLA expenses from Excel file into Neon database.
Matches by last name against members table.

Usage:
  pip install openpyxl psycopg2-binary python-dotenv --break-system-packages
  python import-expenses.py

Place this script in the root of your Stormont project.
Reads DATABASE_URL from .env.local
"""

import os
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv

# Load .env.local
load_dotenv('.env.local')

DATABASE_URL = os.getenv('DATABASE_URL')
if not DATABASE_URL:
    print('ERROR: DATABASE_URL not found in .env.local')
    sys.exit(1)

# Financial year and period — update these when re-running for new data
FINANCIAL_YEAR = '2025-2026'
PERIOD = 'April 2025 - March 2026'
CURRENT_MANDATE = '2022-2027'
EXCEL_PATH = 'expenses.xlsx'

# Read Excel
wb = load_workbook(EXCEL_PATH, read_only=True)
ws = wb.active

rows = []
institutional_rows = []
for row in ws.iter_rows(values_only=True):
    # Rows where first column is an integer (the key) are per-MLA rows
    if isinstance(row[0], int) and row[1]:
        rows.append({
            'key': row[0],
            'name_raw': row[1],
            'constituency_office': float(row[2] or 0),
            'other_expenses': float(row[3] or 0),
            'allowances': float(row[4] or 0),
            'staff_costs': float(row[5] or 0),
            'total': float(row[6] or 0),
        })
    # Rows with no key but a named category and a total are Assembly-wide
    # line items not attributable to an individual MLA (e.g. "Disability &
    # Security Measures Costs"). The summary "Total" row itself is excluded.
    elif row[0] is None and row[1] and str(row[1]).strip().lower() != 'total' and row[6]:
        institutional_rows.append({
            'category': str(row[1]).strip(),
            'amount': float(row[6] or 0),
        })

print(f'Read {len(rows)} MLA rows and {len(institutional_rows)} institutional row(s) from Excel')
if institutional_rows:
    for r in institutional_rows:
        print(f'  Institutional: {r["category"]} = {r["amount"]}')

# Connect to database
import psycopg2
conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

# Fetch all members — build lookup by last name
cur.execute("SELECT person_id, full_name FROM members")
members = cur.fetchall()

# Post-nominal letters that can trail a full_name and must not be read as a surname
POST_NOMINALS = r'(OBE|MBE|CBE|MC|QC|KC|DL|JP)'

# Build last name lookup
# full_name is like "Mr John Smith" or "Ms Claire Sugden" or "Mr Doug Beattie MC"
def get_last_name(full_name):
    # Strip prefix
    import re
    name = re.sub(r'^(Mr|Mrs|Miss|Ms|Dr|Lord|Lady|Sir|Rt Hon Sir|Rt Hon)\s+', '', full_name, flags=re.IGNORECASE).strip()
    # Strip trailing post-nominal letters
    name = re.sub(r'\s+' + POST_NOMINALS + r'(\s+' + POST_NOMINALS + r')*$', '', name, flags=re.IGNORECASE).strip()
    # Last word is last name
    parts = name.split()
    return parts[-1].lower() if parts else ''

def get_first_name(full_name):
    import re
    name = re.sub(r'^(Mr|Mrs|Miss|Ms|Dr|Lord|Lady|Sir|Rt Hon Sir|Rt Hon)\s+', '', full_name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+' + POST_NOMINALS + r'(\s+' + POST_NOMINALS + r')*$', '', name, flags=re.IGNORECASE).strip()
    parts = name.split()
    return parts[0].lower() if len(parts) > 1 else ''

# Excel row names that don't resolve via the last-name lookup above,
# mapped directly to the matching person_id in the members table.
NAME_OVERRIDES = {
    'fleming-archibald, caoimhe': '5800',   # DB: Dr Caoimhe Archibald
    'little pengelly, mary': '5333',        # DB: Mrs Emma Little-Pengelly (typo'd first name in source file)
    'ni chuilin, caral': '215',             # DB: Ms Car�l N� Chuil�n (diacritics stored differently)
    'nichol, kate': '8072',                 # DB: Ms Kate Nicholl (surname spelling differs by one letter)
}

# Build lookup: last_name -> list of (person_id, full_name)
last_name_lookup = {}
for person_id, full_name in members:
    ln = get_last_name(full_name)
    if ln not in last_name_lookup:
        last_name_lookup[ln] = []
    last_name_lookup[ln].append((person_id, full_name))

# Parse Excel name format: "Surname, Firstname" or "Surname Firstname"
def parse_excel_name(name_raw):
    name_raw = name_raw.strip()
    if ',' in name_raw:
        parts = name_raw.split(',', 1)
        last = parts[0].strip().lower()
        first = parts[1].strip().lower().split()[0] if parts[1].strip() else ''
    else:
        parts = name_raw.split()
        last = parts[0].strip().lower()
        first = parts[1].strip().lower() if len(parts) > 1 else ''
    return last, first

# Match each Excel row to a person_id
matched = []
unmatched = []

for row in rows:
    override_id = NAME_OVERRIDES.get(row['name_raw'].strip().lower())
    if override_id:
        matched.append((override_id, row))
        continue

    last, first = parse_excel_name(row['name_raw'])
    candidates = last_name_lookup.get(last, [])

    if len(candidates) == 1:
        # Unique last name match
        person_id = candidates[0][0]
        matched.append((person_id, row))
    elif len(candidates) > 1:
        # Multiple people with same last name — match on first name too
        found = None
        for person_id, full_name in candidates:
            db_first = get_first_name(full_name).lower()
            if db_first.startswith(first[:3]):  # first 3 chars of first name
                found = person_id
                break
        if found:
            matched.append((found, row))
        else:
            unmatched.append(row['name_raw'])
    else:
        unmatched.append(row['name_raw'])

print(f'Matched: {len(matched)}, Unmatched: {len(unmatched)}')

if unmatched:
    print('UNMATCHED NAMES — these will not be imported:')
    for name in unmatched:
        print(f'  - {name}')
    print()

# Insert into database
inserted = 0
for person_id, row in matched:
    cur.execute("""
        INSERT INTO expenses (
            person_id, financial_year, period,
            constituency_office, other_expenses, allowances,
            staff_costs, total, mandate, updated_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        ON CONFLICT (person_id, financial_year)
        DO UPDATE SET
            period = EXCLUDED.period,
            constituency_office = EXCLUDED.constituency_office,
            other_expenses = EXCLUDED.other_expenses,
            allowances = EXCLUDED.allowances,
            staff_costs = EXCLUDED.staff_costs,
            total = EXCLUDED.total,
            mandate = EXCLUDED.mandate,
            updated_at = NOW()
    """, (
        person_id,
        FINANCIAL_YEAR,
        PERIOD,
        row['constituency_office'],
        row['other_expenses'],
        row['allowances'],
        row['staff_costs'],
        row['total'],
        CURRENT_MANDATE,
    ))
    inserted += 1

# Insert institutional (non-MLA) line items
institutional_inserted = 0
for row in institutional_rows:
    cur.execute("""
        INSERT INTO institutional_expenses (
            financial_year, category, amount, mandate, updated_at
        ) VALUES (%s, %s, %s, %s, NOW())
        ON CONFLICT (financial_year, category)
        DO UPDATE SET
            amount = EXCLUDED.amount,
            mandate = EXCLUDED.mandate,
            updated_at = NOW()
    """, (
        FINANCIAL_YEAR,
        row['category'],
        row['amount'],
        CURRENT_MANDATE,
    ))
    institutional_inserted += 1

conn.commit()
cur.close()
conn.close()

print(f'Successfully inserted/updated {inserted} expense records')
print(f'Successfully inserted/updated {institutional_inserted} institutional expense record(s)')
print(f'Financial year: {FINANCIAL_YEAR}')
print(f'Period: {PERIOD}')
